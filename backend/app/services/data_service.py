from __future__ import annotations

from functools import lru_cache
from typing import Any
import math
import re
import unicodedata

from openpyxl import load_workbook

from app.core.config import EXCEL_PATH

CATEGORY_FIELDS = [
    "Edificación colapsada",
    "Edificación con daño",
    "Vivienda afectada",
    "Hospital o centro de salud",
    "Escuela o colegio",
    "Patrimonio o templo",
    "Servicio público",
    "Vía afectada",
    "Deslizamiento",
    "Albergue o punto de atención",
    "Punto de ayuda o acopio",
    "Novedad o noticia",
    "Toque de queda o restricción",
    "Otro",
    "Puente",
    "Incendio",
    "Saqueo",
    "Robo",
]


def _clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
    return value


def _number(value: Any) -> float:
    value = _clean(value)
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(".", "").replace(",", ".") if re.fullmatch(r"\d{1,3}(\.\d{3})*,\d+", str(value)) else str(value)
    try:
        return float(text)
    except (TypeError, ValueError):
        return 0


def _int(value: Any) -> int:
    return int(round(_number(value)))


def _norm(value: Any) -> str:
    text = str(_clean(value) or "")
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", text).strip().lower()


def _rows(sheet):
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(v).strip() if v is not None else "" for v in next(iterator)]
    for values in iterator:
        yield {headers[i]: _clean(values[i]) if i < len(values) else None for i in range(len(headers))}


@lru_cache(maxsize=1)
def load_data() -> dict[str, Any]:
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    # "Análisis general" has title rows; its real table header is row 3.
    general_sheet = wb["Análisis general"]
    general_rows = []
    for values in general_sheet.iter_rows(min_row=4, values_only=True):
        department = _clean(values[1] if len(values) > 1 else None)
        if not department or _norm(department) == "total general":
            continue
        general_rows.append({
            "Departamento": department,
            "N.º de municipios": _clean(values[2] if len(values) > 2 else None),
            "Puntos totales": _clean(values[3] if len(values) > 3 else None),
            "Indice de Pobreza - 2025": _clean(values[4] if len(values) > 4 else None),
        })

    # Consolidated headline indicators are a separate block in the same sheet.
    official_indicators = {}
    for values in general_sheet.iter_rows(min_row=4, values_only=True):
        label = _clean(values[9] if len(values) > 9 else None)
        value = _clean(values[10] if len(values) > 10 else None)
        if label and value is not None:
            official_indicators[_norm(label)] = value
    operational_indicators = {}
    for values in general_sheet.iter_rows(min_row=4, values_only=True):
        label = _clean(values[6] if len(values) > 6 else None)
        value = _clean(values[7] if len(values) > 7 else None)
        if label and value is not None:
            operational_indicators[_norm(label)] = value
    municipalities_raw = list(_rows(wb["Municipios completos"]))
    news_raw = list(_rows(wb["Noticias"]))

    ipm_2025: dict[str, float] = {}
    for row in general_rows:
        dept = _clean(row.get("Departamento"))
        if dept:
            ipm_2025[_norm(dept)] = _number(row.get("Indice de Pobreza - 2025"))

    news_index: dict[tuple[str, str], dict[str, Any]] = {}
    for row in news_raw:
        dept = _clean(row.get("Departamento"))
        muni = _clean(row.get("Municipio"))
        if not dept or not muni:
            continue
        news_index[(_norm(dept), _norm(muni))] = {
            "fallecidos": _int(row.get("Fallecidos ") or row.get("Fallecidos")),
            "heridos": _int(row.get("Heridos")),
            "noticia_1": _clean(row.get("Noticia 1")),
            "noticia_2": _clean(row.get("Noticia 2")),
        }

    municipalities: list[dict[str, Any]] = []
    for row in municipalities_raw:
        divipola = _clean(row.get("Código DIVIPOLA"))
        muni = _clean(row.get("Municipio"))
        dept = _clean(row.get("Departamento"))
        if not divipola or not muni or not dept or _norm(muni) == "total":
            continue
        divipola = str(divipola).split(".")[0].zfill(5)
        news = news_index.get((_norm(dept), _norm(muni)), {})
        item = {
            "divipola": divipola,
            "municipio": muni,
            "departamento": dept,
            "clasificacion": _clean(row.get("Clasificación")),
            "criticidad": _clean(row.get("Clasificación")),
            "puntos": _int(row.get("Puntos")),
            "afectados_familia": _int(row.get("Afectados Familia")),
            "afectados_personas": _int(row.get("Afectados Personas")),
            "ipm_2018": _number(row.get("Indice de Pobreza - 2018")),
            "ipm_2025_departamento": ipm_2025.get(_norm(dept), 0),
            "ayudas": _clean(row.get("ayudas")),
            "tipo_ayuda": _clean(row.get("tipo de ayuda")),
            "descripcion": _clean(row.get("Descripción / Tipo de afectación")),
            "fallecidos": news.get("fallecidos", 0),
            "heridos": news.get("heridos", 0),
            "noticia_1": news.get("noticia_1"),
            "noticia_2": news.get("noticia_2"),
            "categorias": {field: _int(row.get(field)) for field in CATEGORY_FIELDS},
        }
        municipalities.append(item)

    departments = sorted({m["departamento"] for m in municipalities})
    return {
        "municipalities": municipalities,
        "departments": departments,
        "categories": CATEGORY_FIELDS,
        "official_summary": {
            "departamentos": _int(official_indicators.get("departamentos afectados")),
            "municipios": _int(official_indicators.get("municipios afectados")),
            "puntos": _int(official_indicators.get("puntos totales")),
            "fallecidos": _int(operational_indicators.get("fallecidos")),
            "heridos": _int(operational_indicators.get("heridos")),
            "viviendas_destruidas": _int(operational_indicators.get("viviendas destruidas")),
            "viviendas_averiadas": _int(operational_indicators.get("viviendas averiadas")),
            "viviendas_afectadas": _int(official_indicators.get("viviendas afectadas")),
            "centros_salud": _int(operational_indicators.get("centros de salud")),
            "centros_educativos": _int(operational_indicators.get("centros educativos")),
            "vias_afectadas": _int(operational_indicators.get("vias afectadas")),
            "magnitud_epicentro": operational_indicators.get("magnitud y epicentro"),
        },
    }


def filter_municipalities(department: str | None = None, municipality: str | None = None, category: str | None = None):
    data = load_data()["municipalities"]
    result = data
    if department:
        result = [m for m in result if _norm(m["departamento"]) == _norm(department)]
    if municipality:
        result = [m for m in result if _norm(m["municipio"]) == _norm(municipality)]
    if category and category in CATEGORY_FIELDS:
        result = [m for m in result if m["categorias"].get(category, 0) > 0]
    return result


def make_summary(rows: list[dict[str, Any]]) -> dict[str, Any]:
    categories = {field: sum(m["categorias"][field] for m in rows) for field in CATEGORY_FIELDS}
    departments = sorted({m["departamento"] for m in rows})
    return {
        "departamentos": len(departments),
        "municipios": len(rows),
        "puntos": sum(m["puntos"] for m in rows),
        "afectados_familia": sum(m["afectados_familia"] for m in rows),
        "afectados_personas": sum(m["afectados_personas"] for m in rows),
        "heridos": sum(m["heridos"] for m in rows),
        "fallecidos": sum(m["fallecidos"] for m in rows),
        "categorias": categories,
    }
